"""
ChainFlow — sample_data/create_template.py
Generates inventory_template.xlsx for use with POST /inventory/upload/excel.

Usage:
    cd chainflow
    python sample_data/create_template.py

Output: sample_data/inventory_template.xlsx

Re-run any time to regenerate the file (existing file is overwritten).
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Template data
# Eight realistic rows for Harpreet Hosiery Works.
# Unit costs are in INR.
#
# ELAS-BRAID-1IN is intentionally below its reorder threshold (qty 320 < 400)
# so that GET /inventory/alerts?tenant_id=1 returns a visible result on a
# fresh upload — demonstrating the alert system without any manual data entry.
# ──────────────────────────────────────────────────────────────────────────────

HEADERS = [
    "sku_code",
    "name",
    "category",
    "unit",
    "current_quantity",
    "reorder_threshold",
    "reorder_quantity",
    "unit_cost",
]

ROWS = [
    # sku_code          name                           category       unit     qty    thresh  reorder  cost
    ("NYL-FIT-12MM",    "Nylon Fitting 12mm",          "Components",  "units", 4800,  1000,   5000,    1.20),
    ("ELAS-YARN-20MM",  "Elastic Yarn 20mm Width",     "Raw Material","metres",850,   500,    2000,    3.50),
    ("NYL-THREAD-40",   "Nylon Thread 40 Denier",      "Raw Material","kg",    120,   50,     200,     280.00),
    ("PKG-POLY-BAG-M",  "Polybag Medium 12x18 inch",   "Packaging",   "units", 9500,  2000,   10000,   0.80),
    ("DRAW-CORD-3MM",   "Drawcord 3mm Round",          "Components",  "metres",650,   300,    1500,    2.20),
    ("ELAS-BRAID-1IN",  "Braided Elastic 1 inch",      "Raw Material","metres",320,   400,    2000,    6.00),  # ← below threshold
    ("PKG-LABEL-STD",   "Standard Brand Label",        "Packaging",   "units", 3200,  1000,   5000,    0.45),
    ("BUTTON-PLST-15MM","Plastic Button 15mm Assorted","Components",  "units", 7500,  2000,   8000,    0.30),
]

INSTRUCTIONS = [
    ("Column",            "Description",                                            "Example"),
    ("sku_code",          "Unique code for this item within your business",          "NYL-FIT-12MM"),
    ("name",              "Human-readable name as it appears in Tally",              "Nylon Fitting 12mm"),
    ("category",          "One of: Raw Material, Components, Packaging",             "Components"),
    ("unit",              "Unit of measure: kg, metres, units, etc.",                "units"),
    ("current_quantity",  "Stock on hand right now",                                 "4800"),
    ("reorder_threshold", "Alert fires when quantity drops below this level",        "1000"),
    ("reorder_quantity",  "How much to order when restocking",                       "5000"),
    ("unit_cost",         "Cost per unit in INR (used for stock valuation)",         "1.20"),
    ("",                  "",                                                        ""),
    ("Alert behaviour",   "When current_quantity < reorder_threshold, the item",     ""),
    ("",                  "appears in GET /inventory/alerts sorted by urgency.",     ""),
    ("",                  "The row ELAS-BRAID-1IN (qty 320, threshold 400) is",     ""),
    ("",                  "intentionally below threshold to demonstrate this.",      ""),
    ("",                  "",                                                        ""),
    ("Upload",            "POST /inventory/upload/excel with this file as the body.", ""),
    ("",                  "Existing SKUs (matched by sku_code) will be updated.",    ""),
    ("",                  "New SKU codes will be created automatically.",            ""),
]

# ──────────────────────────────────────────────────────────────────────────────
# Styles
# ──────────────────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="D9E1F2")       # light blue
HEADER_FONT  = Font(bold=True)
ALERT_FILL   = PatternFill("solid", fgColor="FCE4D6")        # light orange — below threshold row
INSTR_HEADER_FILL = PatternFill("solid", fgColor="E2EFDA")   # light green for instructions header


def _autofit_columns(ws) -> None:
    """Set each column width to fit its longest cell value (approximate)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ──────────────────────────────────────────────────────────────────────────────
# Build workbook
# ──────────────────────────────────────────────────────────────────────────────

def create_template(output_path: Path) -> None:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Inventory data ───────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "Inventory"

    # Header row
    ws_data.append(HEADERS)
    for col_num, _ in enumerate(HEADERS, start=1):
        cell = ws_data.cell(row=1, column=col_num)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Comment on A1 so the purpose is visible on open
    from openpyxl.comments import Comment
    ws_data["A1"].comment = Comment(
        "Upload this file to POST /inventory/upload/excel\n"
        "See the Instructions sheet for column details.",
        author="ChainFlow",
    )

    # Data rows
    for row_idx, row_data in enumerate(ROWS, start=2):
        ws_data.append(list(row_data))
        # Highlight the intentionally-below-threshold row
        if row_data[0] == "ELAS-BRAID-1IN":
            for col_num in range(1, len(HEADERS) + 1):
                ws_data.cell(row=row_idx, column=col_num).fill = ALERT_FILL

    # Freeze header row and auto-fit
    ws_data.freeze_panes = "A2"
    _autofit_columns(ws_data)

    # ── Sheet 2: Instructions ─────────────────────────────────────────────────
    ws_instr = wb.create_sheet("Instructions")
    ws_instr.column_dimensions["A"].width = 22
    ws_instr.column_dimensions["B"].width = 58
    ws_instr.column_dimensions["C"].width = 26

    for row_idx, (col_name, description, example) in enumerate(INSTRUCTIONS, start=1):
        ws_instr.cell(row=row_idx, column=1, value=col_name)
        ws_instr.cell(row=row_idx, column=2, value=description)
        ws_instr.cell(row=row_idx, column=3, value=example)

    # Style the instructions header row
    for col_num in range(1, 4):
        cell = ws_instr.cell(row=1, column=col_num)
        cell.fill = INSTR_HEADER_FILL
        cell.font = HEADER_FONT

    # Wrap text in description column
    for row in ws_instr.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # ── Save ──────────────────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Template written to: {output_path.resolve()}")
    print(f"  {len(ROWS)} data rows across {len(set(r[2] for r in ROWS))} categories")
    print(f"  ELAS-BRAID-1IN is below threshold — will appear in /inventory/alerts")


if __name__ == "__main__":
    output = Path(__file__).parent / "inventory_template.xlsx"
    create_template(output)
